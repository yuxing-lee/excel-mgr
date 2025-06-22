from flask import Flask, request, jsonify, render_template
import openpyxl
import os
import sys
from datetime import datetime, date, timedelta

BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
EXCEL_FILE = os.path.join(os.getcwd(), 'data.xlsx')


def load_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Option", "Date"])
        wb.save(EXCEL_FILE)
    return openpyxl.load_workbook(EXCEL_FILE)


def read_data():
    wb = load_workbook()
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value = row[3]
        if isinstance(date_value, (datetime, date)):
            date_value = date_value.strftime('%Y-%m-%d')
        data.append({"id": row[0], "name": row[1], "option": row[2], "date": date_value})
    return data


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/data')
def list_data():
    search = request.args.get('search', '').lower()
    option_filter = request.args.get('option')
    data = read_data()
    if search:
        data = [row for row in data if search in str(row['id']).lower() or search in str(row['name']).lower()]
    if option_filter:
        data = [row for row in data if str(row['option']) == option_filter]
    return jsonify(data)


@app.route('/api/soon')
def soon():
    days = request.args.get('days', default=7, type=int)
    data = read_data()
    now = date.today()
    upcoming = []
    for row in data:
        date_str = row.get('date')
        if not date_str:
            continue
        try:
            due = datetime.strptime(str(date_str), '%Y-%m-%d').date()
        except ValueError:
            continue
        if now <= due <= now + timedelta(days=days):
            upcoming.append(row)
    return jsonify(upcoming)


@app.route('/api/add', methods=['POST'])
def add_row():
    id_ = request.json.get('id')
    name = request.json.get('name')
    option = request.json.get('option')
    date_str = request.json.get('date')
    if not id_ or not name or option is None:
        return jsonify({"error": "id, name and option required"}), 400
    date_val = None
    if date_str:
        try:
            date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "invalid date"}), 400
    wb = load_workbook()
    ws = wb.active
    ws.append([id_, name, option, date_val])
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/update', methods=['POST'])
def update_row():
    id_ = request.json.get('id')
    name = request.json.get('name')
    option = request.json.get('option')
    date_str = request.json.get('date')
    if not id_ or name is None or option is None:
        return jsonify({"error": "id, name and option required"}), 400
    date_val = None
    if date_str:
        try:
            date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "invalid date"}), 400
    wb = load_workbook()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(id_):
            row[1].value = name
            row[2].value = option
            if date_str is not None:
                row[3].value = date_val
            wb.save(EXCEL_FILE)
            return jsonify({"success": True})
    return jsonify({"error": "ID not found"}), 404


@app.route('/api/delete', methods=['POST'])
def delete_row():
    id_ = request.json.get('id')
    if not id_:
        return jsonify({"error": "id required"}), 400
    wb = load_workbook()
    ws = wb.active
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(id_):
            row_to_delete = idx
            break
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)
        return jsonify({"success": True})
    else:
        return jsonify({"error": "ID not found"}), 404


if __name__ == '__main__':
    app.run(debug=True)
